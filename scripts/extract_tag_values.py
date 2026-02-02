"""
Extract Tag Values Script
==========================
Extracts all unique tag values and their frequencies from a checkpoint file,
clusters similar values, and generates an audit spreadsheet for review.

Purpose:
  After human review of tagged questions, normalize tag values to ensure
  consistency. Examples: "Bone fracture" vs "Bone fractures", "dara" vs
  "daratumumab", "KEYNOTE 189" vs "KEYNOTE-189".

Process:
  1. Extract unique values per field from final_tags
  2. Cluster similar values using case-insensitive matching, singular/plural,
     whitespace/hyphen variants, and fuzzy matching
  3. Output an Excel spreadsheet with one sheet per field group for user review

Usage:
  # Generate audit spreadsheet from checkpoint data:
  python scripts/extract_tag_values.py

  # Specify a different checkpoint:
  python scripts/extract_tag_values.py --checkpoint data/checkpoints/stage2_tagged_multiple_myeloma.json

  # Change output path:
  python scripts/extract_tag_values.py --output data/eval/tag_value_audit.xlsx
"""

import argparse
import json
import logging
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[logging.StreamHandler(sys.stdout)]
)
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# Fields to normalize (free-text/autocomplete) — grouped by category
FIELD_GROUPS = {
    "treatments": ["treatment_1", "treatment_2", "treatment_3", "treatment_4", "treatment_5"],
    "biomarkers": ["biomarker_1", "biomarker_2", "biomarker_3", "biomarker_4", "biomarker_5"],
    "trials": ["trial_1", "trial_2", "trial_3", "trial_4", "trial_5"],
    "drug_class": ["drug_class_1", "drug_class_2", "drug_class_3"],
    "drug_target": ["drug_target_1", "drug_target_2", "drug_target_3"],
    "prior_therapy": ["prior_therapy_1", "prior_therapy_2", "prior_therapy_3"],
    "metastatic_site": ["metastatic_site_1", "metastatic_site_2", "metastatic_site_3"],
    "symptoms": ["symptom_1", "symptom_2", "symptom_3"],
    "toxicity": ["toxicity_type_1", "toxicity_type_2", "toxicity_type_3",
                 "toxicity_type_4", "toxicity_type_5", "toxicity_organ"],
    "efficacy": ["efficacy_endpoint_1", "efficacy_endpoint_2", "efficacy_endpoint_3"],
    "disease_type": ["disease_type_1", "disease_type_2"],
    "other": ["resistance_mechanism", "special_population_1", "special_population_2",
              "disease_specific_factor"],
}

# Fields to SKIP (closed dropdown / constrained values)
SKIP_FIELDS = {
    "topic", "disease_stage", "treatment_line", "performance_status",
    "cme_outcome_level", "data_response_type", "stem_type", "lead_in_type",
    "answer_format", "answer_length_pattern", "distractor_homogeneity",
    "toxicity_grade", "outcome_context", "clinical_benefit",
    "guideline_source_1", "guideline_source_2", "evidence_type",
    "treatment_eligibility", "age_group", "fitness_status",
    "flaw_absolute_terms", "flaw_grammatical_cue", "flaw_implausible_distractor",
    "flaw_clang_association", "flaw_convergence_vulnerability", "flaw_double_negative",
    "is_oncology", "answer_option_count", "correct_answer_position",
}


# ---------------------------------------------------------------------------
# Clustering Logic
# ---------------------------------------------------------------------------

def normalize_for_comparison(s: str) -> str:
    """Normalize a string for comparison: lowercase, collapse whitespace/hyphens."""
    s = s.lower().strip()
    s = re.sub(r'[\s\-_]+', ' ', s)  # Collapse whitespace, hyphens, underscores
    return s


def singularize(word: str) -> str:
    """Simple singularization for common English plural patterns."""
    if word.endswith('ies') and len(word) > 4:
        return word[:-3] + 'y'
    if word.endswith('es') and len(word) > 3:
        # Handle cases like "fractures" -> "fracture"
        base = word[:-2]
        if base.endswith(('s', 'x', 'z', 'ch', 'sh')):
            return base
        return word[:-1]  # "responses" -> "response"
    if word.endswith('s') and not word.endswith('ss') and len(word) > 3:
        return word[:-1]
    return word


def are_singular_plural_match(a: str, b: str) -> bool:
    """Check if two strings are singular/plural variants of each other."""
    a_norm = normalize_for_comparison(a)
    b_norm = normalize_for_comparison(b)

    if a_norm == b_norm:
        return True

    # Try singularizing each word in the string
    a_words = a_norm.split()
    b_words = b_norm.split()

    if len(a_words) != len(b_words):
        return False

    # Check if they match when last word is singularized
    a_singular = ' '.join(a_words[:-1] + [singularize(a_words[-1])])
    b_singular = ' '.join(b_words[:-1] + [singularize(b_words[-1])])

    return a_singular == b_singular or a_singular == b_norm or a_norm == b_singular


def levenshtein_distance(s1: str, s2: str) -> int:
    """Compute Levenshtein edit distance between two strings."""
    if len(s1) < len(s2):
        return levenshtein_distance(s2, s1)

    if len(s2) == 0:
        return len(s1)

    previous_row = list(range(len(s2) + 1))
    for i, c1 in enumerate(s1):
        current_row = [i + 1]
        for j, c2 in enumerate(s2):
            insertions = previous_row[j + 1] + 1
            deletions = current_row[j] + 1
            substitutions = previous_row[j] + (c1 != c2)
            current_row.append(min(insertions, deletions, substitutions))
        previous_row = current_row

    return previous_row[-1]


def similarity_ratio(a: str, b: str) -> float:
    """Compute similarity ratio between two strings (0-1)."""
    a_norm = normalize_for_comparison(a)
    b_norm = normalize_for_comparison(b)
    if a_norm == b_norm:
        return 1.0
    max_len = max(len(a_norm), len(b_norm))
    if max_len == 0:
        return 1.0
    dist = levenshtein_distance(a_norm, b_norm)
    return 1.0 - (dist / max_len)


def cluster_values(values_with_counts: list[tuple[str, int]],
                   fuzzy_threshold: float = 0.90) -> list[dict]:
    """Cluster similar values and suggest canonical forms.

    Args:
        values_with_counts: List of (value, count) tuples
        fuzzy_threshold: Minimum similarity ratio for fuzzy matching

    Returns:
        List of dicts with keys: value, count, suggested_canonical, action, match_type
    """
    if not values_with_counts:
        return []

    # Sort by count descending (most frequent first)
    sorted_values = sorted(values_with_counts, key=lambda x: -x[1])

    results = []
    processed = set()

    for i, (val, count) in enumerate(sorted_values):
        if val in processed:
            continue

        # This value becomes a potential canonical (most frequent in its cluster)
        cluster = [(val, count, "keep", "canonical")]
        processed.add(val)

        # Find matches for this value
        for j, (other_val, other_count) in enumerate(sorted_values):
            if other_val in processed:
                continue

            match_type = None

            # 1. Case-insensitive exact match
            if normalize_for_comparison(val) == normalize_for_comparison(other_val):
                match_type = "case_match"

            # 2. Singular/plural match
            elif are_singular_plural_match(val, other_val):
                match_type = "plural_match"

            # 3. Whitespace/hyphen variant
            elif normalize_for_comparison(val) == normalize_for_comparison(other_val):
                match_type = "whitespace_match"

            # 4. Fuzzy match
            elif similarity_ratio(val, other_val) >= fuzzy_threshold:
                match_type = "fuzzy_match"

            if match_type:
                cluster.append((other_val, other_count, "merge", match_type))
                processed.add(other_val)

        # Add cluster to results
        canonical = cluster[0][0]  # Most frequent value
        for c_val, c_count, c_action, c_match in cluster:
            results.append({
                "value": c_val,
                "count": c_count,
                "suggested_canonical": canonical,
                "action": c_action,
                "match_type": c_match,
            })

    return results


# ---------------------------------------------------------------------------
# Extraction
# ---------------------------------------------------------------------------

def extract_field_values(data: list, fields: list[str]) -> Counter:
    """Extract all non-null, non-empty values for the given fields."""
    counter = Counter()
    for q in data:
        final_tags = q.get("final_tags", {})
        for field in fields:
            val = final_tags.get(field)
            if val and val.strip():
                counter[val.strip()] += 1
    return counter


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Extract and cluster tag values from checkpoint data for normalization review."
    )
    parser.add_argument(
        "--checkpoint",
        type=str,
        default=str(PROJECT_ROOT / "data" / "checkpoints" / "stage2_tagged_multiple_myeloma.json"),
        help="Path to checkpoint JSON file"
    )
    parser.add_argument(
        "--output",
        type=str,
        default=str(PROJECT_ROOT / "data" / "eval" / "tag_value_audit.xlsx"),
        help="Output path for audit spreadsheet"
    )
    parser.add_argument(
        "--fuzzy-threshold",
        type=float,
        default=0.90,
        help="Minimum similarity ratio for fuzzy matching (default: 0.90)"
    )

    args = parser.parse_args()

    checkpoint_path = Path(args.checkpoint)
    output_path = Path(args.output)

    if not checkpoint_path.exists():
        logger.error(f"Checkpoint file not found: {checkpoint_path}")
        sys.exit(1)

    # Ensure output directory exists
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Load data
    logger.info(f"Loading checkpoint: {checkpoint_path}")
    with open(checkpoint_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    logger.info(f"Loaded {len(data)} questions")

    # Try to import openpyxl for Excel output
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        has_openpyxl = True
    except ImportError:
        has_openpyxl = False
        logger.warning("openpyxl not installed. Falling back to CSV output.")
        output_path = output_path.with_suffix('.csv')

    # Extract and cluster values for each field group
    all_results = {}
    total_values = 0
    total_clusters_with_variants = 0

    for group_name, fields in FIELD_GROUPS.items():
        counter = extract_field_values(data, fields)
        if not counter:
            logger.info(f"  {group_name}: No values found")
            continue

        values_with_counts = list(counter.items())
        clustered = cluster_values(values_with_counts, fuzzy_threshold=args.fuzzy_threshold)

        # Count clusters with variants
        merge_count = sum(1 for r in clustered if r["action"] == "merge")

        all_results[group_name] = clustered
        total_values += len(clustered)
        total_clusters_with_variants += merge_count

        logger.info(f"  {group_name}: {len(counter)} unique values, {merge_count} suggested merges")

    if not all_results:
        logger.info("No tag values found in checkpoint data.")
        return

    # Write output
    if has_openpyxl:
        _write_excel(all_results, output_path)
    else:
        _write_csv(all_results, output_path)

    logger.info(f"\nAudit file written: {output_path}")
    logger.info(f"Total unique values: {total_values}")
    logger.info(f"Suggested merges: {total_clusters_with_variants}")
    logger.info("\nNext steps:")
    logger.info("1. Review the audit file — confirm or override suggested_canonical and action columns")
    logger.info("2. Save as tag_normalization_mappings.xlsx (or .csv)")
    logger.info("3. Run: python scripts/apply_tag_normalization.py --mappings <path>")


def _write_excel(all_results: dict, output_path: Path) -> None:
    """Write results to an Excel workbook with one sheet per field group."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    header_font = Font(bold=True)
    merge_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    canonical_fill = PatternFill(start_color="D5E8D4", end_color="D5E8D4", fill_type="solid")

    for group_name, results in all_results.items():
        ws = wb.create_sheet(title=group_name[:31])  # Excel sheet name limit

        # Headers
        headers = ["value", "count", "suggested_canonical", "action", "match_type"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font

        # Data rows
        for row_idx, row_data in enumerate(
            sorted(results, key=lambda x: (-x["count"], x["value"])), start=2
        ):
            ws.cell(row=row_idx, column=1, value=row_data["value"])
            ws.cell(row=row_idx, column=2, value=row_data["count"])
            ws.cell(row=row_idx, column=3, value=row_data["suggested_canonical"])
            ws.cell(row=row_idx, column=4, value=row_data["action"])
            ws.cell(row=row_idx, column=5, value=row_data["match_type"])

            # Highlight merge rows
            if row_data["action"] == "merge":
                for col in range(1, 6):
                    ws.cell(row=row_idx, column=col).fill = merge_fill
            elif row_data["action"] == "keep":
                for col in range(1, 6):
                    ws.cell(row=row_idx, column=col).fill = canonical_fill

        # Auto-width columns
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 2, 60)

    wb.save(output_path)


def _write_csv(all_results: dict, output_path: Path) -> None:
    """Write results to a single CSV file with a group_name column."""
    import csv

    fieldnames = ["group", "value", "count", "suggested_canonical", "action", "match_type"]
    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for group_name, results in all_results.items():
            for row_data in sorted(results, key=lambda x: (-x["count"], x["value"])):
                writer.writerow({
                    "group": group_name,
                    **row_data,
                })


if __name__ == "__main__":
    main()
