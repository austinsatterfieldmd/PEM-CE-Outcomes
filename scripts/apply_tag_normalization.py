"""
Apply Tag Normalization Script
================================
Applies tag value normalization mappings to checkpoint data, ensuring
consistent terminology across all tagged questions.

Takes a reviewed normalization mappings file (from extract_tag_values.py
output, after user review) and applies the canonical value replacements
to the checkpoint JSON.

Process:
  1. Load normalization mappings (value -> canonical)
  2. For each question in checkpoint, replace matching values
  3. Optionally update normalization_rules.yaml with new aliases
  4. Save checkpoint (with backup)

Usage:
  # Preview changes (dry run):
  python scripts/apply_tag_normalization.py --mappings data/eval/tag_normalization_mappings.xlsx --dry-run

  # Apply normalization:
  python scripts/apply_tag_normalization.py --mappings data/eval/tag_normalization_mappings.xlsx

  # Apply and update normalization_rules.yaml:
  python scripts/apply_tag_normalization.py --mappings data/eval/tag_normalization_mappings.xlsx --update-rules

  # Specify a different checkpoint:
  python scripts/apply_tag_normalization.py --checkpoint data/checkpoints/stage2_tagged_multiple_myeloma.json --mappings data/eval/tag_normalization_mappings.xlsx
"""

import argparse
import csv
import json
import logging
import shutil
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import yaml

PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[logging.StreamHandler(sys.stdout)]
)
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# All normalizable fields (same as extract_tag_values.py)
NORMALIZABLE_FIELDS = [
    "treatment_1", "treatment_2", "treatment_3", "treatment_4", "treatment_5",
    "biomarker_1", "biomarker_2", "biomarker_3", "biomarker_4", "biomarker_5",
    "trial_1", "trial_2", "trial_3", "trial_4", "trial_5",
    "drug_class_1", "drug_class_2", "drug_class_3",
    "drug_target_1", "drug_target_2", "drug_target_3",
    "prior_therapy_1", "prior_therapy_2", "prior_therapy_3",
    "metastatic_site_1", "metastatic_site_2", "metastatic_site_3",
    "symptom_1", "symptom_2", "symptom_3",
    "toxicity_type_1", "toxicity_type_2", "toxicity_type_3",
    "toxicity_type_4", "toxicity_type_5", "toxicity_organ",
    "efficacy_endpoint_1", "efficacy_endpoint_2", "efficacy_endpoint_3",
    "disease_type_1", "disease_type_2",
    "resistance_mechanism", "special_population_1", "special_population_2",
    "disease_specific_factor",
]

TAG_CONTAINERS = ["final_tags", "gpt_tags", "claude_tags", "gemini_tags"]

# Field name to normalization_rules.yaml section mapping
FIELD_TO_RULE_SECTION = {
    "treatment": "treatment_aliases",
    "biomarker": "biomarker_aliases",
    "trial": "trial_aliases",
    "drug_class": "drug_class_aliases",
    "drug_target": "drug_target_aliases",
    "prior_therapy": "prior_therapy_aliases",
    "metastatic_site": "metastatic_site_aliases",
    "symptom": "symptom_aliases",
    "toxicity_type": "toxicity_aliases",
    "toxicity_organ": "toxicity_aliases",
    "efficacy_endpoint": "efficacy_endpoint_aliases",
    "disease_type": "disease_type_aliases",
    "resistance_mechanism": "other_aliases",
    "special_population": "other_aliases",
    "disease_specific_factor": "other_aliases",
}


# ---------------------------------------------------------------------------
# Load Mappings
# ---------------------------------------------------------------------------

def load_mappings_from_excel(path: Path) -> dict[str, str]:
    """Load normalization mappings from an Excel file.

    Expects columns: value, suggested_canonical, action
    Only rows with action='merge' are used.

    Returns: dict mapping original value -> canonical value
    """
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl is required to read Excel files. Install with: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(path, read_only=True)
    mappings = {}

    for ws in wb.worksheets:
        # Find column indices from header row
        headers = {}
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value:
                headers[cell.value.lower().strip()] = col_idx

        value_col = headers.get("value")
        canonical_col = headers.get("suggested_canonical")
        action_col = headers.get("action")

        if not all([value_col, canonical_col, action_col]):
            logger.warning(f"Sheet '{ws.title}' missing required columns, skipping")
            continue

        for row in ws.iter_rows(min_row=2):
            action = row[action_col - 1].value
            if action and action.strip().lower() == "merge":
                original = row[value_col - 1].value
                canonical = row[canonical_col - 1].value
                if original and canonical and original != canonical:
                    mappings[original.strip()] = canonical.strip()

    wb.close()
    return mappings


def load_mappings_from_csv(path: Path) -> dict[str, str]:
    """Load normalization mappings from a CSV file.

    Expects columns: value, suggested_canonical, action
    Only rows with action='merge' are used.
    """
    mappings = {}
    with open(path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            action = (row.get("action") or "").strip().lower()
            if action == "merge":
                original = (row.get("value") or "").strip()
                canonical = (row.get("suggested_canonical") or "").strip()
                if original and canonical and original != canonical:
                    mappings[original] = canonical
    return mappings


def load_mappings(path: Path) -> dict[str, str]:
    """Load mappings from either Excel or CSV based on file extension."""
    if path.suffix.lower() in ('.xlsx', '.xls'):
        return load_mappings_from_excel(path)
    elif path.suffix.lower() == '.csv':
        return load_mappings_from_csv(path)
    else:
        logger.error(f"Unsupported file format: {path.suffix}")
        sys.exit(1)


# ---------------------------------------------------------------------------
# Apply Normalization
# ---------------------------------------------------------------------------

def apply_normalization(data: list, mappings: dict[str, str],
                        dry_run: bool = False) -> tuple[int, int]:
    """Apply normalization mappings to checkpoint data.

    Returns: (questions_modified, total_replacements)
    """
    questions_modified = 0
    total_replacements = 0

    for q in data:
        qid = q.get("question_id", "?")
        q_modified = False

        for container_name in TAG_CONTAINERS:
            container = q.get(container_name)
            if not container or not isinstance(container, dict):
                continue

            for field in NORMALIZABLE_FIELDS:
                val = container.get(field)
                if val and val.strip() in mappings:
                    new_val = mappings[val.strip()]
                    if not dry_run:
                        container[field] = new_val
                    logger.info(
                        f"  {'[DRY RUN] ' if dry_run else ''}"
                        f"Q{qid}.{container_name}.{field}: \"{val}\" -> \"{new_val}\""
                    )
                    total_replacements += 1
                    q_modified = True

        # Also update field_votes
        field_votes = q.get("field_votes", {})
        if field_votes:
            for field in NORMALIZABLE_FIELDS:
                vote = field_votes.get(field)
                if not vote or not isinstance(vote, dict):
                    continue
                for vote_key in ["final_value", "gpt_value", "claude_value", "gemini_value"]:
                    val = vote.get(vote_key)
                    if val and val.strip() in mappings:
                        new_val = mappings[val.strip()]
                        if not dry_run:
                            vote[vote_key] = new_val
                        total_replacements += 1
                        q_modified = True

        if q_modified:
            questions_modified += 1

    return questions_modified, total_replacements


# ---------------------------------------------------------------------------
# Update normalization_rules.yaml
# ---------------------------------------------------------------------------

def get_field_base(field: str) -> str:
    """Get the base field name (without numeric suffix)."""
    # Remove trailing _1, _2, etc.
    parts = field.rsplit('_', 1)
    if len(parts) == 2 and parts[1].isdigit():
        return parts[0]
    return field


def update_normalization_rules(mappings: dict[str, str], rules_path: Path,
                               dry_run: bool = False) -> None:
    """Append new alias mappings to normalization_rules.yaml."""
    # Load existing rules
    if rules_path.exists():
        with open(rules_path, 'r', encoding='utf-8') as f:
            rules = yaml.safe_load(f) or {}
    else:
        rules = {}

    # We can't determine field-specific sections from the mappings alone
    # (a value might appear in multiple fields), so we put them in a
    # generic "aliases" section organized by the canonical value
    new_aliases = rules.get("normalization_aliases", {})

    for original, canonical in mappings.items():
        # Add to the aliases dict: canonical -> list of aliases
        if canonical not in new_aliases:
            new_aliases[canonical] = []
        if original not in new_aliases[canonical]:
            new_aliases[canonical].append(original)

    rules["normalization_aliases"] = new_aliases

    # Add metadata
    rules.setdefault("_metadata", {})
    rules["_metadata"]["last_updated"] = datetime.now().isoformat()
    rules["_metadata"]["total_aliases"] = sum(len(v) for v in new_aliases.values())

    if not dry_run:
        with open(rules_path, 'w', encoding='utf-8') as f:
            yaml.dump(rules, f, default_flow_style=False, allow_unicode=True,
                      sort_keys=False)
        logger.info(f"Updated normalization rules: {rules_path}")
    else:
        logger.info(f"[DRY RUN] Would update normalization rules: {rules_path}")
        logger.info(f"  {len(mappings)} new aliases would be added")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Apply tag value normalization mappings to checkpoint data."
    )
    parser.add_argument(
        "--checkpoint",
        type=str,
        default=str(PROJECT_ROOT / "data" / "checkpoints" / "stage2_tagged_multiple_myeloma.json"),
        help="Path to checkpoint JSON file"
    )
    parser.add_argument(
        "--mappings",
        type=str,
        required=True,
        help="Path to normalization mappings file (Excel or CSV)"
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Preview changes without writing to disk"
    )
    parser.add_argument(
        "--update-rules",
        action="store_true",
        help="Also update config/normalization_rules.yaml with new aliases"
    )

    args = parser.parse_args()

    checkpoint_path = Path(args.checkpoint)
    mappings_path = Path(args.mappings)

    if not checkpoint_path.exists():
        logger.error(f"Checkpoint file not found: {checkpoint_path}")
        sys.exit(1)

    if not mappings_path.exists():
        logger.error(f"Mappings file not found: {mappings_path}")
        sys.exit(1)

    # Load mappings
    logger.info(f"Loading mappings: {mappings_path}")
    mappings = load_mappings(mappings_path)
    logger.info(f"Loaded {len(mappings)} normalization mappings")

    if not mappings:
        logger.info("No merge mappings found. Nothing to do.")
        return

    # Show mappings summary
    logger.info("\nMappings to apply:")
    for original, canonical in sorted(mappings.items()):
        logger.info(f"  \"{original}\" -> \"{canonical}\"")

    # Load checkpoint
    logger.info(f"\nLoading checkpoint: {checkpoint_path}")
    with open(checkpoint_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    logger.info(f"Loaded {len(data)} questions")

    # Apply normalization
    logger.info("\nApplying normalization...")
    questions_modified, total_replacements = apply_normalization(
        data, mappings, dry_run=args.dry_run
    )

    logger.info(f"\nQuestions modified: {questions_modified}")
    logger.info(f"Total replacements: {total_replacements}")

    # Save checkpoint
    if questions_modified > 0 and not args.dry_run:
        # Create backup
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_path = checkpoint_path.with_suffix(f'.pre_normalization_{timestamp}.json')
        shutil.copy2(checkpoint_path, backup_path)
        logger.info(f"Backup created: {backup_path}")

        # Save
        with open(checkpoint_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        logger.info(f"Saved checkpoint: {checkpoint_path}")

    # Update normalization_rules.yaml
    if args.update_rules:
        rules_path = PROJECT_ROOT / "config" / "normalization_rules.yaml"
        update_normalization_rules(mappings, rules_path, dry_run=args.dry_run)

    if args.dry_run:
        logger.info("\n[DRY RUN] No changes written to disk.")
    else:
        logger.info("\nDone!")


if __name__ == "__main__":
    main()
